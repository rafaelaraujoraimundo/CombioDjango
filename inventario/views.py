from django.urls import reverse_lazy
from django.db.models import Sum, Q, Count
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.shortcuts import get_object_or_404, redirect, render
from dashboard.models import BiCentroCusto, BiEstabelecimento, BiFuncionariosCombio
from api_v1.schema import Memory
from inventario.tasks import populate_hardware_data
from .models import (AccountInfo, AcoesProntuario, BIOS, Celular, Computador, ControleFones,
    Controlekit, CPU, Estoque, Hardware, Monitor, ProntuarioCelular, ProntuarioComputador,
    ProntuarioMonitor, Software, Status, Storage, TipoItem)
from .forms import (AcoesProntuarioForm, CelularForm, ComputadorForm, ControleFonesForm,
    ControlekitForm, EstoqueForm, MonitorForm, ProntuarioCelularForm, ProntuarioComputadorForm,
    ProntuarioMonitorForm, StatusForm, TipoItemForm)
from django.contrib.auth.decorators import login_required, permission_required
from django.utils.decorators import method_decorator
from django.contrib import messages
from django.utils.timezone import now
from django.http import HttpResponse
from openpyxl import Workbook


@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_admin_admin', login_url='erro_page'), name='dispatch')
class TipoItemList(ListView):
    model = TipoItem
    queryset = TipoItem.objects.all()
    template_name = 'inventario/tipoitem/tipoitem_list.html'

    def get_context_data(self, **kwargs):
        context = super(TipoItemList, self).get_context_data(**kwargs)
        context['activegroup'] = 'administration'
        context['title'] = 'Tipos de Itens'
        return context

# View para criação de novos tipos de itens
@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_admin_admin', login_url='erro_page'), name='dispatch')
class TipoItemCreate(CreateView):
    model = TipoItem
    form_class = TipoItemForm
    template_name = 'inventario/tipoitem/tipoitem_form.html'
    success_url = reverse_lazy('tipoitem_list')

    def get_context_data(self, **kwargs):
        context = super(TipoItemCreate, self).get_context_data(**kwargs)
        context['activegroup'] = 'administration'
        context['title'] = 'Inclusão de Tipo de Item'
        return context

# View para editar tipos de itens existentes
@login_required(login_url='account_login')
@permission_required('global_permissions.combio_admin_admin', login_url='erro_page')
def TipoItem_edit(request, tipoitem_id):
    activegroup = 'administration'
    title = 'Edição de Tipo de Item'
    context = {'activegroup': activegroup, 'title': title}
    tipoitem = get_object_or_404(TipoItem, pk=tipoitem_id)
    if request.method == "POST":
        form = TipoItemForm(request.POST, instance=tipoitem)
        if form.is_valid():
            form.save()
            return redirect('tipoitem_list')
        else:
            context['form'] = form
    else:
        form = TipoItemForm(instance=tipoitem)
        context['form'] = form
    return render(request, 'inventario/tipoitem/tipoitem_edit.html', context)


@login_required(login_url='account_login')
@permission_required('global_permissions.combio_admin_admin', login_url='erro_page')
def tipoItem_delete(request, tipoitem_id):
    tipoitem = get_object_or_404(TipoItem, pk=tipoitem_id)
    if request.method == "POST":
        tipoitem.delete()
        messages.success(request, f'"{tipoitem.nome}" foi excluído com sucesso.')
        return redirect('tipoitem_list')

    return redirect('tipoitem_list')



class StatusList(ListView):
    model = Status
    queryset = Status.objects.all()
    template_name = 'inventario/status/status_list.html'

    def get_context_data(self, **kwargs):
        context = super(StatusList, self).get_context_data(**kwargs)
        context['activegroup'] = 'administration'
        context['title'] = 'Lista de Status'
        return context

# View para criação de novos Status
@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_admin_admin', login_url='erro_page'), name='dispatch')
class StatusCreate(CreateView):
    model = Status
    form_class = StatusForm
    template_name = 'inventario/status/status_form.html'
    success_url = reverse_lazy('status_list')

    def get_context_data(self, **kwargs):
        context = super(StatusCreate, self).get_context_data(**kwargs)
        context['activegroup'] = 'administration'
        context['title'] = 'Inclusão de Status'
        return context

# View para edição de Status
@login_required(login_url='account_login')
@permission_required('global_permissions.combio_admin_admin', login_url='erro_page')
def status_edit(request, status_id):
    activegroup = 'administration'
    title = 'Edição de Status'
    context = {'activegroup': activegroup, 'title': title}
    status = get_object_or_404(Status, pk=status_id)
    if request.method == "POST":
        form = StatusForm(request.POST, instance=status)
        if form.is_valid():
            form.save()
            return redirect('status_list')
        else:
            context['form'] = form
    else:
        form = StatusForm(instance=status)
        context['form'] = form
    return render(request, 'inventario/status/status_edit.html', context)

# View para exclusão de Status
@login_required(login_url='account_login')
@permission_required('global_permissions.combio_admin_admin', login_url='erro_page')
def status_delete(request, status_id):
    status = get_object_or_404(Status, pk=status_id)
    if request.method == "POST":
        status.delete()
        messages.success(request, f'"{status.nome_status}" foi excluído com sucesso.')
        return redirect('status_list')

    return redirect('status_list')


@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class EstoqueList(ListView):
    model = Estoque
    template_name = 'inventario/estoque/estoque_list.html'
    context_object_name = 'itens_estoque'

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(tipo_item__nome__icontains=search_query) |
                Q(modelo__icontains=search_query) |
                Q(fabricante__icontains=search_query) |
                Q(status__nome_status__icontains=search_query) |
                Q(observacao__icontains=search_query) |
                Q(local__icontains=search_query)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['activegroup'] = 'inventario'
        context['title'] = 'Lista de Itens de Estoque'
        context['search'] = self.request.GET.get('search', '')
        return context
# View para criação de novos itens de estoque
@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class EstoqueCreate(CreateView):
    model = Estoque
    form_class = EstoqueForm
    template_name = 'inventario/estoque/estoque_form.html'
    success_url = reverse_lazy('estoque_list')

    def get_context_data(self, **kwargs):
        context = super(EstoqueCreate, self).get_context_data(**kwargs)
        context['activegroup'] = 'inventario'
        context['title'] = 'Inclusão de Item no Estoque'
        return context

# View para editar itens de estoque existentes
@login_required(login_url='account_login')
@permission_required('global_permissions.combio_inventario', login_url='erro_page')
def estoque_edit(request, estoque_id):
    activegroup = 'inventario'
    title = 'Edição de Item no Estoque'
    context = {'activegroup': activegroup, 'title': title}
    estoque = get_object_or_404(Estoque, pk=estoque_id)
    if request.method == "POST":
        form = EstoqueForm(request.POST, instance=estoque)
        if form.is_valid():
            form.save()
            return redirect('estoque_list')
        else:
            context['form'] = form
    else:
        form = EstoqueForm(instance=estoque)
        context['form'] = form
    return render(request, 'inventario/estoque/estoque_edit.html', context)

# View para exclusão de itens de estoque
@login_required(login_url='account_login')
@permission_required('global_permissions.combio_inventario', login_url='erro_page')
def estoque_delete(request, estoque_id):
    estoque = get_object_or_404(Estoque, pk=estoque_id)
    if request.method == "POST":
        estoque.delete()
        messages.success(request, f'O item "{estoque.modelo}" foi excluído com sucesso.')
        return redirect('estoque_list')

    return redirect('estoque_list')

@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class ControlekitList(ListView):
    model = Controlekit
    template_name = 'inventario/controlekit/controlekit_list.html'
    context_object_name = 'kits'

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(matricula__icontains=search_query) |
                Q(usuario__icontains=search_query) |
                Q(modelo__icontains=search_query) |
                Q(estabelecimento__icontains=search_query) |
                Q(centro_custo__icontains=search_query) |
                Q(serie__icontains=search_query)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Lista de Kits'
        context['activegroup'] = 'inventario'
        context['search'] = self.request.GET.get('search', '')
        return context

# View para criar novo kit
@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class ControlekitCreate(CreateView):
    model = Controlekit
    form_class = ControlekitForm
    template_name = 'inventario/controlekit/controlekit_form.html'
    success_url = reverse_lazy('controlekit_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuarios_list'] = BiFuncionariosCombio.objects.filter(mes=now().month, ano=now().year).values('cdn_funcionario', 'nom_funcionario', 'cdn_estab')
        context['estabelecimentos_list'] = BiEstabelecimento.objects.all().values('estabelecimento', 'sigla_unidade')
        context['centros_custo_list'] = BiCentroCusto.objects.all().values('centrocusto', 'descricaocusto')
        context['activegroup'] = 'inventario'
        context['title'] = 'Criação de Kit'
        return context
    
    def form_valid(self, form):
        # Verifique os dados recebidos e certifique-se de que estão corretos
        print(form.cleaned_data)  # Para debug
        return super().form_valid(form)


# View para editar kit
@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class ControlekitUpdate(UpdateView):
    model = Controlekit
    form_class = ControlekitForm
    template_name = 'inventario/controlekit/controlekit_edit.html'
    success_url = reverse_lazy('controlekit_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Passar os dados necessários para preencher os datalists
        context['usuarios_list'] = BiFuncionariosCombio.objects.filter(mes=now().month, ano=now().year).values('cdn_funcionario', 'nom_funcionario', 'cdn_estab')
        context['estabelecimentos_list'] = BiEstabelecimento.objects.all().values('estabelecimento', 'sigla_unidade')
        context['centros_custo_list'] = BiCentroCusto.objects.all().values('centrocusto', 'descricaocusto')
        context['activegroup'] = 'inventario'
        context['title'] = 'Alteração de Kit'
        return context

# View para excluir kit
@login_required(login_url='account_login')
@permission_required('global_permissions.combio_inventario', login_url='erro_page')
def controlekit_delete(request, controlekit_id):
    controlekit = get_object_or_404(Controlekit, pk=controlekit_id)
    if request.method == "POST":
        controlekit.delete()
        messages.success(request, 'Kit excluído com sucesso.')
        return redirect('controlekit_list')
    return redirect('controlekit_list')

@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
@method_decorator(login_required(login_url='account_login'), name='dispatch')
class ControleFonesList(ListView):
    model = ControleFones
    template_name = 'inventario/controlefones/controlefones_list.html'
    context_object_name = 'fones'

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(matricula__icontains=search_query) |
                Q(usuario__icontains=search_query) |
                Q(modelo__icontains=search_query) |
                Q(estabelecimento__icontains=search_query) |
                Q(centro_custo__icontains=search_query) |
                Q(serie__icontains=search_query)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Lista de Fones'
        context['activegroup'] = 'inventario'
        context['search'] = self.request.GET.get('search', '')
        return context

# View para criar novo fone
@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class ControleFonesCreate(CreateView):
    model = ControleFones
    form_class = ControleFonesForm
    template_name = 'inventario/controlefones/controlefones_form.html'
    success_url = reverse_lazy('controlefones_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuarios_list'] = BiFuncionariosCombio.objects.filter(mes=now().month, ano=now().year).values('cdn_funcionario', 'nom_funcionario', 'cdn_estab')
        context['estabelecimentos_list'] = BiEstabelecimento.objects.all().values('estabelecimento', 'sigla_unidade')
        context['centros_custo_list'] = BiCentroCusto.objects.all().values('centrocusto', 'descricaocusto')
        context['activegroup'] = 'inventario'
        context['title'] = 'Criação de Controle de Fones'
        return context

# View para editar fone
@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class ControleFonesUpdate(UpdateView):
    model = ControleFones
    form_class = ControleFonesForm
    template_name = 'inventario/controlefones/controlefones_edit.html'
    success_url = reverse_lazy('controlefones_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuarios_list'] = BiFuncionariosCombio.objects.filter(mes=now().month, ano=now().year).values('cdn_funcionario', 'nom_funcionario', 'cdn_estab')
        context['estabelecimentos_list'] = BiEstabelecimento.objects.all().values('estabelecimento', 'sigla_unidade')
        context['centros_custo_list'] = BiCentroCusto.objects.all().values('centrocusto', 'descricaocusto')
        context['activegroup'] = 'inventario'
        context['title'] = 'Edição de Controle de Fones'
        return context

# View para excluir fone
@login_required(login_url='account_login')
@permission_required('global_permissions.combio_inventario', login_url='erro_page')
def controlefones_delete(request, controlefones_id):
    controlefones = get_object_or_404(ControleFones, pk=controlefones_id)
    if request.method == "POST":
        controlefones.delete()
        messages.success(request, 'Fone excluído com sucesso.')
        return redirect('controlefones_list')
    return redirect('controlefones_list')



@method_decorator(login_required(login_url='account_login'), name='dispatch')
class CelularList(ListView):
    model = Celular
    template_name = 'inventario/celular/celular_list.html'

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(modelo__icontains=search_query) |
                Q(fabricante__icontains=search_query) |
                Q(numero_serie__icontains=search_query) |
                Q(imei__icontains=search_query) |
                Q(numero_linha__icontains=search_query) |
                Q(usuario__icontains=search_query) |
                Q(estabelecimento__icontains=search_query) |
                Q(centro_custo__icontains=search_query)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super(CelularList, self).get_context_data(**kwargs)
        context['title'] = 'Lista de Celulares'
        context['activegroup'] = 'inventario'
        context['search'] = self.request.GET.get('search', '')
        return context


@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class CelularCreate(CreateView):
    model = Celular
    form_class = CelularForm
    template_name = 'inventario/celular/celular_form.html'
    success_url = reverse_lazy('celular_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Inclusão de Celulares'
        context['activegroup'] = 'inventario'
        
        # Listas para autocomplete
        context['status_list'] = Status.objects.all()
        context['usuarios_list'] = BiFuncionariosCombio.objects.filter(
            mes=now().month, ano=now().year
        ).values('cdn_funcionario', 'nom_funcionario', 'cdn_estab')
        
        context['estabelecimentos_list'] = BiEstabelecimento.objects.all().values(
            'estabelecimento', 'sigla_unidade'
        )
        
        context['centros_custo_list'] = BiCentroCusto.objects.all().values(
            'centrocusto', 'descricaocusto'
        )
        
        return context

@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class CelularUpdate(UpdateView):
    model = Celular
    form_class = CelularForm
    template_name = 'inventario/celular/celular_form.html'
    success_url = reverse_lazy('celular_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Lista de usuários, estabelecimentos, centros de custo e status para preencher os campos de datalist
        context['usuarios_list'] = BiFuncionariosCombio.objects.filter(mes=now().month, ano=now().year).values('cdn_funcionario', 'nom_funcionario', 'cdn_estab')
        context['estabelecimentos_list'] = BiEstabelecimento.objects.all().values('estabelecimento', 'sigla_unidade')
        context['centros_custo_list'] = BiCentroCusto.objects.all().values('centrocusto', 'descricaocusto')
        context['status_list'] = Status.objects.all()  # Para o campo de status
        context['title'] = 'Edição de Celular'
        context['activegroup'] = 'inventario'
        return context


@login_required(login_url='account_login')
@permission_required('global_permissions.combio_inventario', login_url='erro_page')
def celular_delete(request, celular_id):
    celular = get_object_or_404(Celular, pk=celular_id)
    if request.method == "POST":
        celular.delete()
        messages.success(request, 'Celular excluído com sucesso.')
        return redirect('celular_list')
    return redirect('celular_list')

class AcoesProntuarioList(ListView):
    model = AcoesProntuario
    queryset = AcoesProntuario.objects.all()
    template_name = 'inventario/acoesprontuario/acoesprontuario_list.html'

    def get_context_data(self, **kwargs):
        context = super(AcoesProntuarioList, self).get_context_data(**kwargs)
        context['activegroup'] = 'administration'
        context['title'] = 'Lista de Ações de Prontuário'
        return context

# Criação de nova Ação de Prontuário
@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class AcoesProntuarioCreate(CreateView):
    model = AcoesProntuario
    form_class = AcoesProntuarioForm
    template_name = 'inventario/acoesprontuario/acoesprontuario_form.html'
    success_url = reverse_lazy('acoesprontuario_list')

    def get_context_data(self, **kwargs):
        context = super(AcoesProntuarioCreate, self).get_context_data(**kwargs)
        context['activegroup'] = 'administration'
        context['title'] = 'Inclusão de Ação de Prontuário'
        return context

# Edição de Ações de Prontuário
@login_required(login_url='account_login')
@permission_required('global_permissions.combio_inventario', login_url='erro_page')
def acoesprontuario_edit(request, acoesprontuario_id):
    activegroup = 'administration'
    title = 'Edição de Ação de Prontuário'
    context = {'activegroup': activegroup, 'title': title}
    acoesprontuario = get_object_or_404(AcoesProntuario, pk=acoesprontuario_id)
    if request.method == "POST":
        form = AcoesProntuarioForm(request.POST, instance=acoesprontuario)
        if form.is_valid():
            form.save()
            return redirect('acoesprontuario_list')
        else:
            context['form'] = form
    else:
        form = AcoesProntuarioForm(instance=acoesprontuario)
        context['form'] = form
    return render(request, 'inventario/acoesprontuario/acoesprontuario_edit.html', context)

# Exclusão de Ações de Prontuário
@login_required(login_url='account_login')
@permission_required('global_permissions.combio_inventario', login_url='erro_page')
def acoesprontuario_delete(request, acoesprontuario_id):
    acoesprontuario = get_object_or_404(AcoesProntuario, pk=acoesprontuario_id)
    if request.method == "POST":
        acoesprontuario.delete()
        messages.success(request, f'"{acoesprontuario.acao}" foi excluído com sucesso.')
        return redirect('acoesprontuario_list')

    return redirect('acoesprontuario_list')


@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class ProntuarioCelularListView(ListView):
    model = ProntuarioCelular
    template_name = 'inventario/celular/prontuario_celular_list.html'
    context_object_name = 'prontuarios'

    def get_queryset(self):
        self.celular = get_object_or_404(Celular, id=self.kwargs['celular_id'])
        return ProntuarioCelular.objects.filter(celular=self.celular)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['celular'] = self.celular
        context['activegroup'] = 'inventario'
        context['title'] = f'Prontuário de {self.celular.fabricante} - {self.celular.modelo} - IMEI: {self.celular.imei}'
        return context

# View para adicionar um novo registro de ação ao prontuário do celular


@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class ProntuarioCelularCreate(CreateView):
    model = ProntuarioCelular
    form_class = ProntuarioCelularForm
    template_name = 'inventario/celular/prontuario_celular_form.html'

    def form_valid(self, form):
        form.instance.celular_id = self.kwargs['celular_id']
        
        # Verifica se a ação é do tipo 2 (Transferência)
        if form.cleaned_data['acao'].tipo == 2:
            celular = form.instance.celular
            # Atualiza o estabelecimento e o centro de custo do celular
            celular.estabelecimento = form.cleaned_data['unidade_destino']
            celular.centro_custo = form.cleaned_data['local']
            celular.save()  # Salva as alterações no celular

        return super().form_valid(form)

    def form_invalid(self, form):
        # Adiciona mensagem de erro quando o formulário é inválido
        messages.error(self.request, "Houve um erro ao tentar salvar o formulário. Verifique os campos e tente novamente.")
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        celular = get_object_or_404(Celular, pk=self.kwargs['celular_id'])
        context['title'] = 'Edição de Ações no Prontuário'
        context['activegroup'] = 'inventario'
        context['celular'] = celular
        context['usuarios_list'] = BiFuncionariosCombio.objects.filter(mes=now().month, ano=now().year).values('cdn_funcionario', 'nom_funcionario', 'cdn_estab')
        context['estabelecimentos_list'] = BiEstabelecimento.objects.all().values('estabelecimento', 'sigla_unidade')
        context['centros_custo_list'] = BiCentroCusto.objects.all().values('centrocusto', 'descricaocusto')
        return context

    def get_success_url(self):
        return reverse_lazy('prontuario_celular_list', kwargs={'celular_id': self.kwargs['celular_id']})

# View para editar um registro de ação no prontuário do celular
@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class ProntuarioCelularUpdate(UpdateView):
    model = ProntuarioCelular
    form_class = ProntuarioCelularForm
    template_name = 'inventario/celular/prontuario_celular_edit.html'
    
    def get_success_url(self):
        return reverse_lazy('prontuario_celular_list', kwargs={'celular_id': self.object.celular.id})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Incluindo o objeto celular no contexto
        context['celular'] = self.object.celular
        context['title'] = 'Edição de Ações no Prontuário'
        context['activegroup'] = 'inventario'

        # Incluindo as listas necessárias para os datalists
        context['usuarios_list'] = BiFuncionariosCombio.objects.filter(mes=now().month, ano=now().year).values('cdn_funcionario', 'nom_funcionario', 'cdn_estab')
        context['estabelecimentos_list'] = BiEstabelecimento.objects.all().values('estabelecimento', 'sigla_unidade')
        context['centros_custo_list'] = BiCentroCusto.objects.all().values('centrocusto', 'descricaocusto')

        return context

# View para excluir um registro de ação do prontuário do celular
@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class ProntuarioCelularDelete(DeleteView):
    model = ProntuarioCelular
    template_name = 'inventario/celular/prontuario_celular_confirm_delete.html'
    
    def get_success_url(self):
        return reverse_lazy('prontuario_celular_list', kwargs={'celular_id': self.object.celular.id})
    


@method_decorator(login_required(login_url='account_login'), name='dispatch')
class ProntuarioMonitorListView(ListView):
    model = ProntuarioMonitor
    template_name = 'inventario/monitor/prontuario_monitor_list.html'
    context_object_name = 'prontuarios'

    def get_queryset(self):
        self.monitor = get_object_or_404(Monitor, id=self.kwargs['monitor_id'])
        return ProntuarioMonitor.objects.filter(monitor=self.monitor)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['monitor'] = self.monitor
        context['activegroup'] = 'inventario'
        context['title'] = f'Prontuário de {self.monitor.fabricante} - {self.monitor.modelo} - Série: {self.monitor.numero_serie}'
        return context

@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class ProntuarioMonitorCreate(CreateView):
    model = ProntuarioMonitor
    form_class = ProntuarioMonitorForm
    template_name = 'inventario/monitor/prontuario_monitor_form.html'

    def form_valid(self, form):
        form.instance.monitor_id = self.kwargs['monitor_id']
        
        # Verifica se a ação é do tipo 2 (Transferência)
        if form.cleaned_data['acao'].tipo == 2:
            monitor = form.instance.monitor
            # Atualiza o estabelecimento e o centro de custo (local) do monitor
            monitor.estabelecimento = form.cleaned_data['unidade_destino']
            monitor.local = form.cleaned_data['local']
            monitor.localizacao = form.cleaned_data['localizacao_destino']
            monitor.save()  # Salva as alterações no monitor

        return super().form_valid(form)

    def form_invalid(self, form):
        # Adiciona mensagem de erro quando o formulário é inválido
        messages.error(self.request, "Houve um erro ao tentar salvar o formulário. Verifique os campos e tente novamente.")
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        monitor = get_object_or_404(Monitor, pk=self.kwargs['monitor_id'])
        context['monitor'] = monitor
        context['activegroup'] = 'inventario'
        context['title'] = 'Adicionar Ação no Prontuário do Monitor'
        context['usuarios_list'] = BiFuncionariosCombio.objects.filter(mes=now().month, ano=now().year).values('cdn_funcionario', 'nom_funcionario', 'cdn_estab')
        context['estabelecimentos_list'] = BiEstabelecimento.objects.all().values('estabelecimento', 'sigla_unidade')
        context['centros_custo_list'] = BiCentroCusto.objects.all().values('centrocusto', 'descricaocusto')
        return context

    def get_success_url(self):
        return reverse_lazy('prontuario_monitor_list', kwargs={'monitor_id': self.kwargs['monitor_id']})


@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class ProntuarioMonitorUpdate(UpdateView):
    model = ProntuarioMonitor
    form_class = ProntuarioMonitorForm
    template_name = 'inventario/monitor/prontuario_monitor_edit.html'
    
    def form_valid(self, form):
        # Verifica se a ação é do tipo 2 (Transferência)
        if form.cleaned_data['acao'].tipo == 2:
            monitor = form.instance.monitor
            # Atualiza o estabelecimento e o centro de custo do monitor
            monitor.estabelecimento = form.cleaned_data['unidade_destino']
            monitor.local = form.cleaned_data['local']
            monitor.localizacao = form.cleaned_data['localizacao_destino']
            monitor.save()  # Salva as alterações no monitor
        
        return super().form_valid(form)

    def form_invalid(self, form):
        # Adiciona mensagem de erro quando o formulário é inválido
        messages.error(self.request, "Houve um erro ao tentar salvar o formulário. Verifique os campos e tente novamente.")
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Inclui o objeto monitor no contexto
        context['monitor'] = self.object.monitor
        context['title'] = 'Editar Ação do Prontuário'
        context['activegroup'] = 'inventario'

        # Inclui as listas necessárias para os datalists
        context['usuarios_list'] = BiFuncionariosCombio.objects.filter(mes=now().month, ano=now().year).values('cdn_funcionario', 'nom_funcionario', 'cdn_estab')
        context['estabelecimentos_list'] = BiEstabelecimento.objects.all().values('estabelecimento', 'sigla_unidade')
        context['centros_custo_list'] = BiCentroCusto.objects.all().values('centrocusto', 'descricaocusto')
        
        return context

    def get_success_url(self):
        return reverse_lazy('prontuario_monitor_list', kwargs={'monitor_id': self.object.monitor.id})

@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class ProntuarioMonitorDelete(DeleteView):
    model = ProntuarioMonitor
    template_name = 'inventario/monitor/prontuario_monitor_confirm_delete.html'
    
    def get_success_url(self):
        return reverse_lazy('prontuario_monitor_list', kwargs={'monitor_id': self.object.monitor.id})
    


@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
@method_decorator(login_required(login_url='account_login'), name='dispatch')
class MonitorList(ListView):
    model = Monitor
    template_name = 'inventario/monitor/monitor_list.html'
    context_object_name = 'monitores'

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(numero_serie__icontains=search_query) |
                Q(fabricante__icontains=search_query) |
                Q(modelo__icontains=search_query) |
                Q(patrimonio__icontains=search_query) |
                Q(estabelecimento__icontains=search_query) |
                Q(local__icontains=search_query) |
                Q(localizacao__icontains=search_query)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Lista de Monitores'
        context['activegroup'] = 'inventario'
        context['search'] = self.request.GET.get('search', '')
        return context

@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class MonitorCreate(CreateView):
    model = Monitor
    form_class = MonitorForm
    template_name = 'inventario/monitor/monitor_form.html'
    success_url = reverse_lazy('monitor_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Inclusão de Monitores'
        context['activegroup'] = 'inventario'
        
        # Listas para autocomplete
        context['status_list'] = Status.objects.all()
        context['usuarios_list'] = BiFuncionariosCombio.objects.filter(
            mes=now().month, ano=now().year
        ).values('cdn_funcionario', 'nom_funcionario', 'cdn_estab')
        
        context['estabelecimentos_list'] = BiEstabelecimento.objects.all().values(
            'estabelecimento', 'sigla_unidade'
        )
        
        context['centros_custo_list'] = BiCentroCusto.objects.all().values(
            'centrocusto', 'descricaocusto'
        )
        
        return context

@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class MonitorUpdate(UpdateView):
    model = Monitor
    form_class = MonitorForm
    template_name = 'inventario/monitor/monitor_edit.html'
    success_url = reverse_lazy('monitor_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Listas para preencher os campos de datalist
        context['usuarios_list'] = BiFuncionariosCombio.objects.filter(
            mes=now().month, ano=now().year
        ).values('cdn_funcionario', 'nom_funcionario', 'cdn_estab')
        
        context['estabelecimentos_list'] = BiEstabelecimento.objects.all().values(
            'estabelecimento', 'sigla_unidade'
        )
        
        context['centros_custo_list'] = BiCentroCusto.objects.all().values(
            'centrocusto', 'descricaocusto'
        )
        
        context['status_list'] = Status.objects.all()  # Para o campo de status
        
        context['title'] = 'Edição de Monitor'
        context['activegroup'] = 'inventario'
        return context

@login_required(login_url='account_login')
@permission_required('global_permissions.combio_inventario', login_url='erro_page')
def monitor_delete(request, monitor_id):
    monitor = get_object_or_404(Monitor, pk=monitor_id)
    if request.method == "POST":
        monitor.delete()
        messages.success(request, 'Monitor excluído com sucesso.')
        return redirect('monitor_list')
    return redirect('monitor_list')


@method_decorator(login_required, name='dispatch')
class ComputadorList(ListView):
    model = Computador
    template_name = 'inventario/computador/computador_list.html'
    context_object_name = 'computadores'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(patrimonio__icontains=search_query) |
                Q(hostname__icontains=search_query) |
                Q(numero_serie__icontains=search_query) |
                Q(fabricante__icontains=search_query) |
                Q(modelo__icontains=search_query) |
                Q(processador__icontains=search_query) |
                Q(memoria__icontains=search_query) |
                Q(hd__icontains=search_query) |
                Q(usuario__icontains=search_query) |
                Q(centro_custo__icontains=search_query) |
                Q(estabelecimento__icontains=search_query) |
                Q(cargo__icontains=search_query) |
                Q(numero_nota_fiscal__icontains=search_query) |
                Q(fornecedor__icontains=search_query) |
                Q(sistema_operacional__icontains=search_query) 
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Lista de Computadores'
        context['activegroup'] = 'inventario'
        context['search'] = self.request.GET.get('search', '')
        return context


@method_decorator(login_required, name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class ComputadorCreate(CreateView):
    model = Computador
    form_class = ComputadorForm
    template_name = 'inventario/computador/computador_form.html'
    success_url = reverse_lazy('computador_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Inclusão de Computador'
        context['activegroup'] = 'inventario'

        # Obter lista de hardware com os campos necessários e incluir dados da BIOS
        hardware_data = []
        for hardware in Hardware.objects.all():
            # Soma total do HD de cada hardware no InventarioStorage
            total_hd = Storage.objects.filter(hardware_id=hardware.id).aggregate(total_hd=Sum('disk_size'))['total_hd'] or 0
            
            # Dados de BIOS relacionados
            bios = hardware.bios.first()  # Assumimos que há apenas uma entrada BIOS por hardware

            hardware_data.append({
                'id': hardware.id,
                'name': hardware.name,
                'device_id': hardware.device_id,
                'memory': hardware.memory,
                'processor_t': hardware.processor_t,
                'os_name': hardware.os_name,
                'total_hd': total_hd,
                'bios_mmodel': bios.mmanufacturer if bios else '',
                'bios_smodel': bios.smodel if bios else '',
                'bios_ssn': bios.ssn if bios else ''
            })
        
        context['hardware_list'] = hardware_data  # Passa todos os dados de hardware necessários
        context['status_list'] = Status.objects.all()
        context['usuarios_list'] = BiFuncionariosCombio.objects.filter(
            mes=now().month, ano=now().year
        ).values('cdn_funcionario', 'nom_funcionario', 'cdn_estab')
        
        context['estabelecimentos_list'] = BiEstabelecimento.objects.all().values(
            'estabelecimento', 'sigla_unidade'
        )
        
        context['centros_custo_list'] = BiCentroCusto.objects.all().values(
            'centrocusto', 'descricaocusto'
        )

        return context

# View para editar computador existente
@method_decorator(login_required, name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class ComputadorUpdate(UpdateView):
    model = Computador
    form_class = ComputadorForm
    template_name = 'inventario/computador/computador_form.html'
    success_url = reverse_lazy('computador_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Edição de Computador'
        context['activegroup'] = 'inventario'
        hardware_data = []
        for hardware in Hardware.objects.all():
            # Soma total do HD de cada hardware no InventarioStorage
            total_hd = Storage.objects.filter(hardware_id=hardware.id).aggregate(total_hd=Sum('disk_size'))['total_hd'] or 0
            
            # Dados de BIOS relacionados
            bios = hardware.bios.first()  # Assumimos que há apenas uma entrada BIOS por hardware

            hardware_data.append({
                'id': hardware.id,
                'name': hardware.name,
                'device_id': hardware.device_id,
                'memory': hardware.memory,
                'processor_t': hardware.processor_t,
                'os_name': hardware.os_name,
                'total_hd': total_hd,
                'bios_mmodel': bios.mmanufacturer if bios else '',
                'bios_smodel': bios.smodel if bios else '',
                'bios_ssn': bios.ssn if bios else ''
            })
        
        context['hardware_list'] = hardware_data  # Passa todos os dados de hardware necessários
        context['status_list'] = Status.objects.all()
        context['usuarios_list'] = BiFuncionariosCombio.objects.filter(
            mes=now().month, ano=now().year
        ).values('cdn_funcionario', 'nom_funcionario', 'cdn_estab')
        
        context['estabelecimentos_list'] = BiEstabelecimento.objects.all().values(
            'estabelecimento', 'sigla_unidade'
        )
        
        context['centros_custo_list'] = BiCentroCusto.objects.all().values(
            'centrocusto', 'descricaocusto'
        )

        return context

# View para excluir um computador
@login_required
@permission_required('global_permissions.combio_inventario', login_url='erro_page')
def computador_delete(request, pk):
    computador = get_object_or_404(Computador, pk=pk)
    if request.method == "POST":
        computador.delete()
        messages.success(request, 'Computador excluído com sucesso.')
        return redirect('computador_list')
    return render(request, 'inventario/computador/computador_confirm_delete.html', {'computador': computador})

# Views para ProntuarioComputador
@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class ProntuarioComputadorListView(ListView):
    model = ProntuarioComputador
    template_name = 'inventario/computador/details/computador_details.html'
    context_object_name = 'prontuarios'

    def get_queryset(self):
        # Assume que 'computador_id' é passado como uma kwarg na URL
        self.computador = get_object_or_404(Computador, id=self.kwargs['computador_id'])
        return ProntuarioComputador.objects.filter(computador=self.computador)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['computador'] = self.computador

        # Verifica se o computador possui hardware associado antes de acessar os detalhes
        if hasattr(self.computador, 'hardware') and self.computador.hardware:
            context['hardware'] = self.computador.hardware
            context['bios'] = self.computador.hardware.bios.all()
            context['cpus'] = self.computador.hardware.cpus.all()
            context['memories'] = self.computador.hardware.memories.all()
            context['softwares'] = self.computador.hardware.software.all()
            context['storages'] = self.computador.hardware.storages.all()
            context['accountinfo'] = self.computador.hardware.accountinfo.all()
        else:
            # Define contextos vazios caso hardware não esteja associado
            context['hardware'] = None
            context['bios'] = []
            context['cpus'] = []
            context['memories'] = []
            context['softwares'] = []
            context['storages'] = []
            context['accountinfo'] = []

        context['activegroup'] = 'inventario'
        context['title'] = f'Prontuário de {self.computador.modelo} - {self.computador.fabricante} - {self.computador.numero_serie}'
        return context
        
@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class ProntuarioComputadorCreate(CreateView):
    model = ProntuarioComputador
    form_class = ProntuarioComputadorForm
    template_name = 'inventario/computador/prontuario_computador_form.html'

    def form_valid(self, form):
        form.instance.computador_id = self.kwargs['computador_id']
       
        # Atualiza o estabelecimento e o centro de custo do computador se a ação é do tipo 2 (Transferência)
        if form.cleaned_data['acao'].tipo == 2:
            computador = form.instance.computador
            computador.estabelecimento = form.cleaned_data['unidade_destino']
            computador.centro_custo = form.cleaned_data['local']
            computador.save()  # Salva as alterações no computador

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        computador = get_object_or_404(Computador, pk=self.kwargs['computador_id'])
        context.update({
            'title': 'Adicionar Ação ao Prontuário',
            'computador': computador,
            'usuarios_list': BiFuncionariosCombio.objects.all().values('cdn_funcionario', 'nom_funcionario', 'cdn_estab'),
            'estabelecimentos_list': BiEstabelecimento.objects.all().values('estabelecimento', 'sigla_unidade'),
            'centros_custo_list': BiCentroCusto.objects.all().values('centrocusto', 'descricaocusto'),
            'activegroup': 'inventario'
        })
        return context

    def get_success_url(self):
        return reverse_lazy('prontuario_computador_list', kwargs={'computador_id': self.kwargs['computador_id']})

@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class ProntuarioComputadorUpdate(UpdateView):
    model = ProntuarioComputador
    form_class = ProntuarioComputadorForm
    template_name = 'inventario/computador/prontuario_computador_edit.html'

    def form_valid(self, form):
        print("Ação Tipo:", form.cleaned_data['acao'].tipo)
        # Atualiza o estabelecimento e o centro de custo do computador se a ação é do tipo 2 (Transferência)
        if form.cleaned_data['acao'].tipo == 2:
            computador = form.instance.computador
            computador.estabelecimento = form.cleaned_data['unidade_destino']
            computador.centro_custo = form.cleaned_data['local']
            print(computador.centro_custo)
            computador.save()  # Salva as alterações no computador

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        computador = self.object.computador
        context.update({
            'title': 'Editar Ação no Prontuário',
            'computador': computador,
            'usuarios_list': BiFuncionariosCombio.objects.all().values('cdn_funcionario', 'nom_funcionario', 'cdn_estab'),
            'estabelecimentos_list': BiEstabelecimento.objects.all().values('estabelecimento', 'sigla_unidade'),
            'centros_custo_list': BiCentroCusto.objects.all().values('centrocusto', 'descricaocusto'),
            'activegroup': 'inventario'
        })
        return context

    def get_success_url(self):
        return reverse_lazy('prontuario_computador_list', kwargs={'computador_id': self.object.computador.id})

@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class ProntuarioComputadorDelete(DeleteView):
    model = ProntuarioComputador
    template_name = 'inventario/computador/prontuario_computador_confirm_delete.html'

    def get_success_url(self):
        return reverse_lazy('prontuario_computador_list', kwargs={'computador_id': self.object.computador.id})

@method_decorator(login_required(login_url='account_login'), name='dispatch')
@method_decorator(permission_required('global_permissions.combio_inventario', login_url='erro_page'), name='dispatch')
class ComputadorDetailView(DetailView):
    model = Computador
    template_name = 'inventario/computador/details/computador_details.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        computador_id = self.kwargs.get('pk')  # Assumindo que a chave primária é passada como 'pk'

        # Obtenha o objeto Computador com prefetch e select_related para otimizar as consultas ao banco de dados
        computador = Computador.objects.select_related('hardware').prefetch_related(
            'hardware__bios',
            'hardware__cpus',
            'hardware__memories',
            'hardware__software',
            'hardware__storages',
            'hardware__accountinfo'
        ).get(pk=computador_id)

        # Adicionando ao contexto
        context['computador'] = computador
        context['hardware'] = computador.hardware
        context['bios'] = computador.hardware.bios.all()
        context['cpus'] = computador.hardware.cpus.all()
        context['memories'] = computador.hardware.memories.all()
        context['softwares'] = computador.hardware.software.all()
        context['storages'] = computador.hardware.storages.all()
        context['accountinfo'] = computador.hardware.accountinfo.all()
        context['activegroup'] = 'inventario'

        return context


@login_required(login_url='account_login')
def estabelecimento_chart(request):

    data_computadores = Computador.objects.values('estabelecimento').annotate(total=Count('id')).order_by('estabelecimento')
    computadores_estabelecimentos = [data['estabelecimento'] for data in data_computadores]
    computadores_totals = [data['total'] for data in data_computadores]

    # Celulares por estabelecimento
    data_celulares = Celular.objects.values('estabelecimento').annotate(total=Count('id')).order_by('estabelecimento')
    celulares_estabelecimentos = [data['estabelecimento'] for data in data_celulares]
    celulares_totals = [data['total'] for data in data_celulares]

    # Monitores por estabelecimento
    data_monitores = Monitor.objects.values('estabelecimento').annotate(total=Count('id')).order_by('estabelecimento')
    monitores_estabelecimentos = [data['estabelecimento'] for data in data_monitores]
    monitores_totals = [data['total'] for data in data_monitores]

    # Computadores por sistema operacional
    data_sistemas = Computador.objects.values('sistema_operacional').annotate(total=Count('id')).order_by('sistema_operacional')
    sistemas_operacionais = [data['sistema_operacional'] for data in data_sistemas]
    sistemas_totals = [data['total'] for data in data_sistemas]
    
    # Computadores por Status
    data_status = Computador.objects.values('status__nome_status').annotate(total=Count('id')).order_by('status__nome_status')
    status_nomes = [data['status__nome_status'] for data in data_status]
    status_totals = [data['total'] for data in data_status]

        # Celulares por Fabricante
    data_celulares_fabricante = Celular.objects.values('fabricante').annotate(total=Count('id')).order_by('fabricante')
    fabricantes_celular = [data['fabricante'] for data in data_celulares_fabricante]
    fabricantes_celular_totals = [data['total'] for data in data_celulares_fabricante]

    # Celulares por Status
    data_celulares_status = Celular.objects.values('status__nome_status').annotate(total=Count('id')).order_by('status__nome_status')
    status_celular = [data['status__nome_status'] for data in data_celulares_status]
    status_celular_totals = [data['total'] for data in data_celulares_status]

    # Monitores por Fabricante
    data_monitores_fabricante = Monitor.objects.values('fabricante').annotate(total=Count('id')).order_by('fabricante')
    fabricantes_monitor = [data['fabricante'] for data in data_monitores_fabricante]
    fabricantes_monitor_totals = [data['total'] for data in data_monitores_fabricante]

    # Monitores por Status
    data_monitores_status = Monitor.objects.values('status__nome_status').annotate(total=Count('id')).order_by('status__nome_status')
    status_monitor = [data['status__nome_status'] for data in data_monitores_status]
    status_monitor_totals = [data['total'] for data in data_monitores_status]


    context = {
        'computadores_estabelecimentos': computadores_estabelecimentos,
        'computadores_totals': computadores_totals,
        'celulares_estabelecimentos': celulares_estabelecimentos,
        'celulares_totals': celulares_totals,
        'monitores_estabelecimentos': monitores_estabelecimentos,
        'monitores_totals': monitores_totals,
        'sistemas_operacionais': sistemas_operacionais,
        'sistemas_totals': sistemas_totals,
        'status_nomes': status_nomes,
        'status_totals': status_totals,
        'fabricantes_celular': fabricantes_celular,
        'fabricantes_celular_totals': fabricantes_celular_totals,
        'status_celular': status_celular,
        'status_celular_totals': status_celular_totals,
        'fabricantes_monitor': fabricantes_monitor,
        'fabricantes_monitor_totals': fabricantes_monitor_totals,
        'status_monitor': status_monitor,
        'status_monitor_totals': status_monitor_totals,
        'activegroup': 'Dashboard',
        'title': 'Dashboard Ativos'
    }
   
    return render(request, 'inventario/dashboard/principal.html', context)

@login_required(login_url='account_login')
@permission_required('global_permissions.combio_inventario', login_url='erro_page')
def export_computadores_excel(request):
    # Obter o termo de pesquisa do request
    search_query = request.GET.get('search', '')

    # Filtrar os computadores conforme a pesquisa
    if search_query:
        computadores = Computador.objects.filter(
            Q(patrimonio__icontains=search_query) |
            Q(hostname__icontains=search_query) |
            Q(numero_serie__icontains=search_query) |
            Q(fabricante__icontains=search_query) |
            Q(modelo__icontains=search_query) |
            Q(processador__icontains=search_query) |
            Q(memoria__icontains=search_query) |
            Q(hd__icontains=search_query) |
            Q(usuario__icontains=search_query) |
            Q(centro_custo__icontains=search_query) |
            Q(estabelecimento__icontains=search_query) |
            Q(cargo__icontains=search_query) |
            Q(numero_nota_fiscal__icontains=search_query) |
            Q(fornecedor__icontains=search_query) |
            Q(sistema_operacional__icontains=search_query)
        )
    else:
        computadores = Computador.objects.all()

    # Criação de um novo workbook do Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Computadores'

    # Adicionando cabeçalhos para as colunas
    columns = ['Hostname', 'Fabricante', 'Modelo', 'Processador', 'Memória', 'HD', 'Usuário', 'Estabelecimento', 'Centro de Custo']
    ws.append(columns)

    # Preenchendo as linhas com os dados dos computadores
    for computador in computadores:
        ws.append([
            computador.hostname, computador.fabricante, computador.modelo,
            computador.processador, computador.memoria, computador.hd,
            computador.usuario, computador.estabelecimento, computador.centro_custo
        ])

    # Preparando a resposta HTTP para enviar o arquivo
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="computadores.xlsx"'

    # Salvar o arquivo Excel na resposta HTTP
    wb.save(response)

    return response

@login_required(login_url='account_login')
@permission_required('global_permissions.combio_inventario', login_url='erro_page')
def export_estoque_excel(request):
    search_query = request.GET.get('search', '')

    if search_query:
        itens_estoque = Estoque.objects.filter(
            Q(tipo_item__nome__icontains=search_query) |
            Q(modelo__icontains=search_query) |
            Q(fabricante__icontains=search_query) |
            Q(status__nome_status__icontains=search_query) |
            Q(observacao__icontains=search_query) |
            Q(local__icontains=search_query)
        )
    else:
        itens_estoque = Estoque.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Itens de Estoque'

    columns = ['Tipo de Item', 'Modelo', 'Fabricante', 'Status', 'Observação', 'Local']
    ws.append(columns)

    for item in itens_estoque:
        ws.append([
            item.tipo_item.nome, item.modelo, item.fabricante,
            item.status.nome_status, item.observacao, item.local
        ])

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="estoque.xlsx"'
    wb.save(response)

    return response

@login_required(login_url='account_login')
@permission_required('global_permissions.combio_inventario', login_url='erro_page')
def export_controlekit_excel(request):
    search_query = request.GET.get('search', '')

    if search_query:
        kits = Controlekit.objects.filter(
            Q(matricula__icontains=search_query) |
            Q(usuario__icontains=search_query) |
            Q(modelo__icontains=search_query) |
            Q(estabelecimento__icontains=search_query) |
            Q(centro_custo__icontains=search_query) |
            Q(serie__icontains=search_query)
        )
    else:
        kits = Controlekit.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Kits'

    columns = ['Matrícula', 'Usuário', 'Modelo', 'Estabelecimento', 'Centro de Custo', 'Série', 'Data de Entrega', 'Data Final']
    ws.append(columns)

    for kit in kits:
        ws.append([
            kit.matricula, kit.usuario, kit.modelo, kit.estabelecimento,
            kit.centro_custo, kit.serie, kit.data_entrega, kit.data_final
        ])

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="controlekit.xlsx"'
    wb.save(response)

    return response

@login_required(login_url='account_login')
@permission_required('global_permissions.combio_inventario', login_url='erro_page')
def export_controlefones_excel(request):
    search_query = request.GET.get('search', '')

    if search_query:
        fones = ControleFones.objects.filter(
            Q(matricula__icontains=search_query) |
            Q(usuario__icontains=search_query) |
            Q(modelo__icontains=search_query) |
            Q(estabelecimento__icontains=search_query) |
            Q(centro_custo__icontains=search_query) |
            Q(serie__icontains=search_query)
        )
    else:
        fones = ControleFones.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Fones'

    columns = ['Matrícula', 'Usuário', 'Modelo', 'Estabelecimento', 'Centro de Custo', 'Série', 'Data de Entrega', 'Data Final']
    ws.append(columns)

    for fone in fones:
        ws.append([
            fone.matricula, fone.usuario, fone.modelo, fone.estabelecimento,
            fone.centro_custo, fone.serie, fone.data_entrega, fone.data_final
        ])

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="controlefones.xlsx"'
    wb.save(response)

    return response

@login_required(login_url='account_login')
@permission_required('global_permissions.combio_inventario', login_url='erro_page')
def export_monitor_excel(request):
    search_query = request.GET.get('search', '')

    if search_query:
        monitores = Monitor.objects.filter(
            Q(numero_serie__icontains=search_query) |
            Q(fabricante__icontains=search_query) |
            Q(modelo__icontains=search_query) |
            Q(patrimonio__icontains=search_query) |
            Q(estabelecimento__icontains=search_query) |
            Q(local__icontains=search_query) |
            Q(localizacao__icontains=search_query)
        )
    else:
        monitores = Monitor.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Monitores'

    columns = ['Número de Série', 'Fabricante', 'Modelo', 'Patrimônio', 'Estabelecimento', 'Local', 'Localização']
    ws.append(columns)

    for monitor in monitores:
        ws.append([
            monitor.numero_serie, monitor.fabricante, monitor.modelo,
            monitor.patrimonio, monitor.estabelecimento, monitor.local,
            monitor.localizacao
        ])

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="monitores.xlsx"'
    wb.save(response)

    return response

@login_required(login_url='account_login')
@permission_required('global_permissions.combio_inventario', login_url='erro_page')
def export_celular_excel(request):
    search_query = request.GET.get('search', '')

    if search_query:
        celulares = Celular.objects.filter(
            Q(modelo__icontains=search_query) |
            Q(fabricante__icontains=search_query) |
            Q(numero_serie__icontains=search_query) |
            Q(imei__icontains=search_query) |
            Q(numero_linha__icontains=search_query) |
            Q(usuario__icontains=search_query) |
            Q(status__nome_status__icontains=search_query) |
            Q(estabelecimento__icontains=search_query) |
            Q(centro_custo__icontains=search_query)
        )
    else:
        celulares = Celular.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Celulares'

    columns = ['Modelo', 'Fabricante', 'Número de Série', 'IMEI', 'Número da Linha', 'Usuário', 'Status', 'Estabelecimento', 'Centro de Custo']
    ws.append(columns)

    for celular in celulares:
        ws.append([
            celular.modelo, celular.fabricante, celular.numero_serie,
            celular.imei, celular.numero_linha, celular.usuario,
            celular.status.nome_status, celular.estabelecimento, celular.centro_custo
        ])

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="celulares.xlsx"'
    wb.save(response)

    return response


